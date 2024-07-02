import streamlit as st
import pymupdf  # PyMuPDF
import pandas as pd
import re
import openpyxl
from io import BytesIO
import base64

# Function to extract text from PDF using PyMuPDF
def extract_text_from_pdf(file):
    text = ""
    with pymupdf.open(stream=file.read(), filetype="pdf") as pdf:
        for page in pdf:
            text += page.get_text()
    return text

# Parsing functions for different formats
def parse_costco(text):
    po_number = re.search(r'Order #:\s*(\d+)', text)
    po_total = re.search(r'Purchase Order Total:\s*([\d,]+\.\d{2})', text) or re.search(r'Merchandise Total\s*\d+\s*([\d,]+\.\d{2})', text)
    units = re.search(r'Total Qty:\s*([\d,]+\.\d{2})', text) or re.search(r'Merchandise Total\s*(\d+)', text)
    ship_date_match = re.search(r'Requested Ship Date:\s*([\d/]+)', text) or re.search(r'Requested Ship Date:\s*\n\s*([\d/]+)', text)
    ship_date = ship_date_match.group(1).strip() if ship_date_match else "N/A"
    region_location = re.search(r'(\b\w{2}\b)\s*COSTCO REGION', text)
    dfi_match = re.search(r'DFI% -\s*([\d.]+)', text)
    dfi = float(dfi_match.group(1).strip()) if dfi_match else "N/A"
    frt_match = re.search(r'FRT% -\s*([\d.]+)', text)
    frt = float(frt_match.group(1).strip()) if frt_match else "N/A"
    sku_match = re.search(r'(\d+)\s+Item Detail:', text)
    item_description_match = re.search(r'Item Detail:\s*(.*?)(?=\s+T\d+|\s+Unit Price:|$)', text, re.DOTALL)

    return {
        "PO_NUM": int(po_number.group(1)) if po_number else "N/A",  # Convert to number
        "PO_DOLLARS": float(po_total.group(1).replace(',', '')) if po_total else "N/A",
        "UNITS": int(float(units.group(1))) if units else "N/A",  # Convert to integer
        "SHIP_DATE": ship_date,
        "REGION": region_location.group(1).strip() if region_location else "N/A",
        "DFI": dfi,
        "FRT": frt,
        "SKU": int(sku_match.group(1).strip()) if sku_match else "N/A",  # Convert to number
        "ITEM_DESCRIPTION": item_description_match.group(1).strip().replace('\n', ' ') if item_description_match else "N/A"
    }

# Function to detect and use the appropriate parsing function
def detect_and_parse(text, parser):
    return parser(text)

# Function to validate format based on text content
def validate_format(text, format_option):
    if format_option == "Costco PO":
        return bool(re.search(r'Order #:\s*\d+', text))
    return False

# Function to provide a download link for the file
def get_table_download_link(output, file_name):
    b64 = base64.b64encode(output.getvalue()).decode()
    return f'''
    <a href="data:file/xlsx;base64,{b64}" download="{file_name}">
        <button class="download-button">Download appended Excel file</button>
    </a>
    '''

# Function to append or update data in the Excel sheet
def append_or_update_excel(wb, df_new_data, column_order):
    sheet = wb.active  # Assuming data goes into the first sheet

    # Ensure DataFrame columns are in the correct order
    df_new_data = df_new_data[column_order]

    # Check if the sheet headers are in the correct order
    headers = [cell.value for cell in sheet[1]]
    if headers != column_order:
        for idx, col_name in enumerate(column_order, start=1):
            sheet.cell(row=1, column=idx, value=col_name)

    # Get existing PO #s and their row indices
    existing_po_nums = {cell.value: cell.row for cell in sheet['A'] if cell.value is not None and cell.row > 1}

    # Update existing rows or append new rows
    for _, row in df_new_data.iterrows():
        if row["PO_NUM"] in existing_po_nums:
            row_idx = existing_po_nums[row["PO_NUM"]]
            for idx, col_name in enumerate(column_order, start=1):
                # Update cell only if new data is not "N/A"
                if row[col_name] != "N/A":
                    sheet.cell(row=row_idx, column=idx, value=row[col_name])
        else:
            new_row = [row[col] for col in column_order]
            sheet.append(new_row)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# Function to read uploaded Excel or CSV file
def read_uploaded_file(file):
    if file.name.endswith('.xlsx'):
        df = pd.read_excel(file)
    elif file.name.endswith('.csv'):
        df = pd.read_csv(file)
    else:
        st.error("Unsupported file type. Please upload an Excel (.xlsx) or CSV (.csv) file.")
        return None
    return df

st.title("Costco Magic Spoon PO PDF Manager ")

# Dropdown to select the format
format_option = st.selectbox("Select the format to read:", ("Costco PO",))

# File uploader widget for PDFs
st.markdown("### Upload PDFs")
uploaded_pdfs = st.file_uploader("Choose PDF files", type="pdf", accept_multiple_files=True)

# File uploader widget for Excel or CSV
st.markdown("### Upload Excel or CSV File")
uploaded_file = st.file_uploader("Choose an Excel or CSV file", type=["xlsx", "csv"])

if uploaded_pdfs and uploaded_file:
    st.success("PDFs and Excel/CSV file uploaded successfully!")

    # Load the existing Excel or CSV file to get existing PO numbers
    df_existing = read_uploaded_file(uploaded_file)
    
    if df_existing is not None:
        # Define the correct order of columns from the uploaded file
        column_order = df_existing.columns.tolist()
        
        # Extract column headers from the existing sheet
        existing_po_nums = df_existing["PO_NUM"].tolist()

        all_parsed_data = []
        duplicate_po_nums = []

        for uploaded_pdf in uploaded_pdfs:
            # Extract text from the uploaded PDF using PyMuPDF
            extracted_text = extract_text_from_pdf(uploaded_pdf)

            # Validate the format of the uploaded PDF
            if not validate_format(extracted_text, format_option):
                st.error(f"The uploaded PDF {uploaded_pdf.name} does not match the selected format ({format_option}). Please select the correct format.")
            else:
                # Choose the parsing function based on the selected format
                parser = parse_costco

                # Parse the extracted text using the selected parsing function
                parsed_data = detect_and_parse(extracted_text, parser)
                if parsed_data["PO_NUM"] not in existing_po_nums:
                    all_parsed_data.append(parsed_data)
                else:
                    duplicate_po_nums.append(parsed_data["PO_NUM"])

        if duplicate_po_nums:
            st.warning(f"The following PO #s already exist in the file: {', '.join(map(str, duplicate_po_nums))}.")

        if all_parsed_data:
            # Create a DataFrame for the parsed data
            df_new_data = pd.DataFrame(all_parsed_data)

            # Ensure all necessary columns are present
            for col in column_order:
                if col not in df_new_data.columns:
                    df_new_data[col] = "N/A"

            # Ensure DataFrame columns are in the correct order
            df_new_data = df_new_data[column_order]

            # Create a form for user input for missing columns
            st.markdown("### Enter missing information")
            for i, row in df_new_data.iterrows():
                with st.expander(f"PO #{row['PO_NUM']}"):
                    df_new_data.at[i, 'ROTATION'] = st.number_input(f"Rotation for PO #{row['PO_NUM']}", key=f"rotation_{row['PO_NUM']}", step=1, format="%d") or "N/A"
                    df_new_data.at[i, 'CARRIER'] = st.text_input(f"Carrier for PO #{row['PO_NUM']}", key=f"carrier_{row['PO_NUM']}") or "N/A"
                    df_new_data.at[i, 'TYPE'] = st.text_input(f"Type for PO #{row['PO_NUM']}", key=f"type_{row['PO_NUM']}") or "N/A"
                    df_new_data.at[i, 'CLIENT'] = st.text_input(f"Client for PO #{row['PO_NUM']}", key=f"client_{row['PO_NUM']}") or "N/A"

            # Display the DataFrame as a table using st.table
            st.markdown("### Extracted Data Table")
            st.table(df_new_data)

            if st.button("Append Data to Excel"):
                # If the uploaded file is an Excel file, append or update data in the existing sheet
                if uploaded_file.name.endswith('.xlsx'):
                    wb = openpyxl.load_workbook(uploaded_file)
                    output = append_or_update_excel(wb, df_new_data, column_order)
                # If the uploaded file is a CSV file, append or update data in the existing CSV
                else:
                    df_combined = pd.concat([df_existing, df_new_data], ignore_index=True)
                    df_combined.drop_duplicates(subset=["PO_NUM"], keep='last', inplace=True)

                    output = BytesIO()
                    df_combined.to_csv(output, index=False)
                    output.seek(0)

                st.success("Data appended or updated successfully in the existing file!")

                if uploaded_file.name.endswith('.xlsx'):
                    updated_xls = pd.ExcelFile(BytesIO(output.getvalue()), engine='openpyxl')
                    updated_df_excel = pd.read_excel(updated_xls, sheet_name=updated_xls.sheet_names[0])
                    st.markdown("### Updated Data in Sheet")
                    st.dataframe(updated_df_excel)
                else:
                    updated_df_csv = pd.read_csv(output)
                    st.markdown("### Updated Data in CSV")
                    st.dataframe(updated_df_csv)

                # Provide download link
                st.markdown(
                    get_table_download_link(output, "updated_file.xlsx" if uploaded_file.name.endswith('.xlsx') else "updated_file.csv"),
                    unsafe_allow_html=True
                )

# Styling for better visual separation and aesthetics
st.markdown(
    """
    <style>
    .css-1aumxhk {
        margin-bottom: 2rem;
    }
    .css-1d391kg {
        padding: 2rem 1.5rem;
        border: 1px solid #ddd;
        border-radius: 5px;
        background-color: #f9f9f9;
    }
    .css-12ttj6m {
        font-size: 1.5rem;
        font-weight: bold;
        margin-bottom: 1rem;
    }
    .css-1gbgaq7 {
        margin-top: 2rem;
    }
    .download-button {
        display: inline-block;
        padding: 10px 20px;
        font-size: 16px;
        cursor: pointer;
        text-align: center;
        text-decoration: none;
        outline: none;
        color: #fff;
        background-color: #007bff;
        border: none;
        border-radius: 5px;
    }
    .download-button:hover {background-color: #0056b3}
    .download-button:active {
        background-color: #0056b3;
    }
    .download-button:disabled {
        background-color: #cccccc;
        color: #666666;
        cursor: not-allowed;
    }
    </style>
    """,
    unsafe_allow_html=True
)

